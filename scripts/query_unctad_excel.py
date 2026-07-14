#!/usr/bin/env python3
"""
query_unctad_excel.py — filter the local UNCTAD ISDS Navigator full-data Excel
snapshot and emit structured, citation-ready results with a mandatory
data-freshness footer.

Part of the isds-research skill. Compliance posture: the Excel is
UNCTAD's intended public data product; filtering happens locally, for
non-commercial research; nothing is republished. This script never fetches
case pages in bulk — its only network action is an optional, throttled
(max 1/day) freshness check that reads the "Updated as of" date from one
server-rendered Navigator page.

Usage examples:
  python query_unctad_excel.py --respondent Argentina --status "favour of investor"
  python query_unctad_excel.py --treaty "Energy Charter" --count-only
  python query_unctad_excel.py --breach-found "Umbrella clause" --json
  python query_unctad_excel.py --case "CMS" --full
  python query_unctad_excel.py --list-values STATUS
  python query_unctad_excel.py --needs-live-check --respondent Spain
"""

import argparse
import glob
import json
import os
import re
import sys
import urllib.parse
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl is required (pip install openpyxl).")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "..", "data")
CACHE_PATH = os.path.join(SCRIPT_DIR, ".freshness-cache.json")

# Last state observed by a human/agent session; used when the live check
# is unavailable. Update when a newer observation is made.
LAST_KNOWN = {"updated_as_of": "31/12/2025", "total": 1463, "observed_on": "2026-07-02"}

FRESHNESS_URL = (
    "https://investmentpolicy.unctad.org/investment-dispute-settlement/cases/68/cms-v-argentina"
)

# Canonical column keys -> exact header text in the 31/12/2023 release.
EXPECTED_HEADERS = {
    "NO": "NO.",
    "YEAR": "YEAR OF INITIATION",
    "SHORT": "SHORT CASE NAME",
    "FULL": "FULL CASE NAME",
    "IIA": "APPLICABLE IIA",
    "RULES": "ARBITRAL RULES",
    "INSTITUTION": "ADMINISTERING INSTITUTION",
    "STATUS": "STATUS/OUTCOME OF ORIGINAL PROCEEDINGS",
    "RESPONDENT": "RESPONDENT STATE",
    "HOME": "HOME STATE OF INVESTOR",
    "SECTOR": "ECONOMIC SECTOR",
    "SUBSECTOR": "ECONOMIC SUBSECTOR",
    "SUMMARY": "SUMMARY OF THE DISPUTE",
    "INVESTMENT": "DETAILS OF INVESTMENT",
    "ARBITRATORS": "ARBITRATORS",
    "DECISIONS": "DECISIONS",
    "OPINIONS": "INDIVIDUAL OPINIONS DETAILS",
    "CLAIMED": "AMOUNT CLAIMED (EXPRESSED IN MILLIONS)",
    "AWARDED": "AMOUNT AWARDED (OR SETTLED FOR) (EXPRESSED IN MILLIONS)",
    "ALLEGED": "IIA BREACHES ALLEGED",
    "FOUND": "IIA BREACHES FOUND",
    "FO_TYPE": "FOLLOW-ON PROCEEDING TYPE",
    "FO_STATUS": "FOLLOW-ON PROCEEDING STATUS",
    "FO_DECISIONS": "FOLLOW-ON DECISIONS",
    "FO_OPINIONS": "FOLLOW-ON INDIVIDUAL OPINIONS",
    "COMMITTEE": "ICSID ANNULMENT COMMITTEE MEMBERS",
    "ITALAW": "LINK TO ITALAW'S CASE PAGE",
    "SOURCES": "LINKS TO BACKGROUND SOURCES",
}
DERIVED_FIELDS = ["ICSID_NO", "LAST_DECISION", "LIVE_CHECK"]
DEFAULT_FIELDS = ["NO", "YEAR", "SHORT", "ICSID_NO", "STATUS", "RESPONDENT", "HOME", "LIVE_CHECK"]

MONTHS = {m.lower(): i for i, m in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"], 1)}

ICSID_NO_RE = re.compile(r"ICSID Case No\.?\s*([A-Z]+(?:\s*\(AF\))?/\d{2}/\d+)")
ICSID_LINK_RE = re.compile(r"CaseNo=([A-Za-z0-9%/()\.]+)")
DECISION_DATE_RE = re.compile(r"dated\s+(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})")

# A concluded case is flagged for live verification if its last recorded
# decision is on/after this date (recent activity => plausibly more since
# the snapshot). Aligned with the routing taxonomy in SKILL.md.
RECENT_DECISION_CUTOFF = date(2021, 1, 1)


DATA_MISSING_MSG = """DATA_MISSING
The UNCTAD full-data Excel was not found in {data_dir} (and no --data was given).

WHY YOU MUST DOWNLOAD IT YOURSELF: UNCTAD's Terms of Use permit personal,
non-commercial use of its ISDS data but bar redistribution — so this tool does
not ship the dataset or fetch it for you. Each user obtains their own copy
directly from UNCTAD (free, no registration). One download; reused thereafter.

GET IT:
  1. Check for the newest full-data release:
     https://investmentpolicy.unctad.org/publications/1303/investment-dispute-settlement-navigator-full-isds-data-release-as-of-31-12-2023-in-excel-format-
  2. Latest known direct link (31/12/2023 snapshot):
     https://investmentpolicy.unctad.org/uploaded-files/document/UNCTAD-ISDS-Navigator-data-set-31December2023.xlsx

THEN: place the .xlsx in the skill's data/ folder. In a Claude chat, simply
upload the downloaded file and ask Claude to save it into the skill's data/
folder — Claude will reuse it in later sessions without re-uploading.

(Award retrieval via fetch_icsid_award.py works without this file; only
case-enumeration questions need it.)"""


def find_data_file(explicit):
    if explicit:
        if not os.path.exists(explicit):
            sys.exit(f"ERROR: data file not found: {explicit}")
        return explicit
    candidates = sorted(glob.glob(os.path.join(DATA_DIR, "*.xlsx")))
    if not candidates:
        print(DATA_MISSING_MSG.format(data_dir=os.path.normpath(DATA_DIR)))
        sys.exit(2)
    if len(candidates) > 1:
        sys.exit("ERROR: multiple .xlsx files in data/; pass --data PATH:\n  "
                 + "\n  ".join(candidates))
    return candidates[0]


def snapshot_date_from(path, wb):
    # Prefer the embedded disclaimer sheet; fall back to the filename.
    for ws in wb.worksheets:
        if "disclaimer" in ws.title.lower():
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if v and "Updated as of" in str(v):
                        m = re.search(r"Updated as of\s+(.+?)\s*$", str(v))
                        if m:
                            return m.group(1).strip()
    m = re.search(r"(\d{1,2})(\w+)(\d{4})", os.path.basename(path))
    return m.group(0) if m else "unknown"


def load_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    snap = snapshot_date_from(path, wb)
    ws = None
    for cand in wb.worksheets:
        if "disclaimer" not in cand.title.lower():
            ws = cand
            break
    if ws is None:
        sys.exit("ERROR: could not find the data sheet.")
    header_map, rows = None, []
    for row in ws.iter_rows(values_only=True):
        vals = ["" if v is None else str(v).strip() for v in row]
        if header_map is None:
            if "NO." in vals and "SHORT CASE NAME" in vals:
                header_map = {}
                norm = [v.upper().strip() for v in vals]
                for key, header in EXPECTED_HEADERS.items():
                    if header.upper() in norm:
                        header_map[key] = norm.index(header.upper())
                missing = set(EXPECTED_HEADERS) - set(header_map)
                if missing:
                    print(f"WARNING: headers not found (schema drift?): {sorted(missing)}",
                          file=sys.stderr)
            continue
        try:
            int(float(vals[header_map["NO"]]))
        except (ValueError, KeyError):
            continue
        rows.append({key: vals[idx] if idx < len(vals) else ""
                     for key, idx in header_map.items()})
    if header_map is None:
        sys.exit("ERROR: header row not found — is this the UNCTAD full-data Excel?")
    return snap, rows


def icsid_no(rec):
    m = ICSID_NO_RE.search(rec.get("FULL", ""))
    if m:
        return re.sub(r"\s+", "", m.group(1))
    m = ICSID_LINK_RE.search(rec.get("SOURCES", ""))
    if m:
        return urllib.parse.unquote(m.group(1)).strip("/.")
    return ""


def last_decision(rec):
    best = None
    for col in ("DECISIONS", "FO_DECISIONS"):
        for d, mon, y in DECISION_DATE_RE.findall(rec.get(col, "")):
            mi = MONTHS.get(mon.lower())
            if not mi:
                continue
            try:
                dt = date(int(y), mi, int(d))
            except ValueError:
                continue
            if best is None or dt > best:
                best = dt
    return best


def live_check_flag(rec):
    reasons = []
    st = rec.get("STATUS", "").lower()
    if "pending" in st or "data not available" in st:
        reasons.append("status pending/unknown at snapshot")
    if "pending" in rec.get("FO_STATUS", "").lower():
        reasons.append("follow-on proceeding pending at snapshot")
    ld = last_decision(rec)
    if ld and ld >= RECENT_DECISION_CUTOFF and not reasons:
        reasons.append(f"recent activity (last decision {ld.isoformat()})")
    return "; ".join(reasons)


def enrich(rec):
    rec["ICSID_NO"] = icsid_no(rec)
    ld = last_decision(rec)
    rec["LAST_DECISION"] = ld.isoformat() if ld else ""
    rec["LIVE_CHECK"] = live_check_flag(rec)
    return rec


def match(rec, args):
    def has(key, needle):
        return needle.lower() in rec.get(key, "").lower()

    checks = [
        ("respondent", "RESPONDENT"), ("home_state", "HOME"), ("treaty", "IIA"),
        ("rules", "RULES"), ("institution", "INSTITUTION"), ("status", "STATUS"),
        ("sector", "SECTOR"), ("subsector", "SUBSECTOR"),
        ("breach_alleged", "ALLEGED"), ("breach_found", "FOUND"),
        ("arbitrator", "ARBITRATORS"), ("committee", "COMMITTEE"),
    ]
    for arg_name, key in checks:
        v = getattr(args, arg_name)
        if v and not has(key, v):
            return False
    if args.follow_on and not (has("FO_TYPE", args.follow_on) or has("FO_STATUS", args.follow_on)):
        return False
    if args.case:
        n = args.case.lower()
        if getattr(args, "case_exact", False):
            if not (n == rec["SHORT"].lower() or n == rec["FULL"].lower()
                    or n == rec["ICSID_NO"].lower()):
                return False
        elif not (n in rec["SHORT"].lower() or n in rec["FULL"].lower()
                  or n in rec["ICSID_NO"].lower()):
            return False
    if args.text:
        if not any(args.text.lower() in v.lower() for v in rec.values()):
            return False
    if args.year:
        m = re.fullmatch(r"(\d{4})(?:-(\d{4}))?", args.year)
        if not m:
            sys.exit("ERROR: --year must be YYYY or YYYY-YYYY")
        lo, hi = int(m.group(1)), int(m.group(2) or m.group(1))
        try:
            y = int(float(rec["YEAR"]))
        except ValueError:
            return False
        if not (lo <= y <= hi):
            return False
    if args.needs_live_check and not rec["LIVE_CHECK"]:
        return False
    return True


def _last_known():
    """LAST_KNOWN fallback, with a staleness warning once the observation ages."""
    prov = f"NOT verified now — last known observation {LAST_KNOWN['observed_on']}"
    try:
        age_d = (datetime.now() - datetime.fromisoformat(LAST_KNOWN["observed_on"])).days
        if age_d > 90:
            prov += (f"; STALE — that observation is {age_d} days old: treat the live-Navigator"
                     " state as UNKNOWN and verify at the Navigator before relying on it")
    except (ValueError, KeyError):
        pass
    return LAST_KNOWN["updated_as_of"], LAST_KNOWN["total"], prov


def freshness(no_network):
    """Return (navigator_updated_as_of, total, provenance)."""
    now = datetime.now()
    if os.path.exists(CACHE_PATH):
        try:
            c = json.load(open(CACHE_PATH))
            age_h = (now - datetime.fromisoformat(c["checked_at"])).total_seconds() / 3600
            if age_h < 24:
                return c["updated_as_of"], c.get("total"), f"live check cached {c['checked_at'][:16]}"
        except Exception:
            pass
    if no_network:
        return _last_known()
    try:
        import urllib.request
        req = urllib.request.Request(FRESHNESS_URL, headers={
            "User-Agent": "isds-research-skill/1.0.0 (non-commercial research; freshness check, max 1/day)"})
        html = urllib.request.urlopen(req, timeout=15).read().decode("utf-8", "replace")
        m_date = re.search(r"Updated as of\s*([0-3]?\d/[01]?\d/\d{4})", html)
        m_total = re.search(r"status=1000[^>]*>\s*([\d,]+)", html)
        if m_date:
            info = {"checked_at": now.isoformat(), "updated_as_of": m_date.group(1),
                    "total": int(m_total.group(1).replace(",", "")) if m_total else None}
            try:
                json.dump(info, open(CACHE_PATH, "w"))
            except OSError:
                pass
            return info["updated_as_of"], info["total"], "verified live just now"
    except Exception as e:
        print(f"NOTE: freshness check failed ({e.__class__.__name__}); using last known.",
              file=sys.stderr)
    return _last_known()


def footer(snap, n_total, args):
    upd, total, prov = freshness(args.no_freshness)
    lag = ""
    if total:
        lag = f" ({total - n_total:+d} cases vs. this snapshot)"
    lines = [
        "=" * 12 + " DATA FRESHNESS (include in any answer) " + "=" * 12,
        f"SNAPSHOT: UNCTAD full-data Excel as of {snap} — {n_total} known treaty-based ISDS cases.",
        "SCOPE: treaty-based, publicly known cases only; per UNCTAD the data 'cannot be deemed exhaustive'.",
        f"LIVE NAVIGATOR: updated as of {upd}{lag} — {prov}. The Navigator itself refreshes ~biannually;",
        "  truly current status must come from the institution's live page (ICSID/PCA).",
        f"GAP: cases initiated after {snap}, and post-snapshot developments (status changes, new decisions,",
        "  follow-on proceedings), are NOT in these results. Verify rows flagged LIVE_CHECK before relying",
        "  on status-sensitive fields.",
        "CITE: UNCTAD, Investment Dispute Settlement Navigator: full data release as of 31/12/2023 (excel",
        "  format), https://investmentpolicy.unctad.org/investment-dispute-settlement",
    ]
    return "\n".join(lines)


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--data", help="path to the UNCTAD Excel (default: sole .xlsx in ../data)")
    for flag in ("respondent", "home-state", "treaty", "rules", "institution", "status",
                 "sector", "subsector", "breach-alleged", "breach-found", "arbitrator",
                 "committee", "follow-on", "text", "year"):
        p.add_argument(f"--{flag}")
    p.add_argument("--case", help="case name or ICSID number. SUBSTRING match — may over-match "
                   "similarly named cases (a shared claimant name matches every case it appears "
                   "in); add --case-exact for whole-field equality")
    p.add_argument("--case-exact", action="store_true",
                   help="with --case: exact (case-insensitive) match against the short case "
                   "name, full case name, or ICSID case number")
    p.add_argument("--needs-live-check", action="store_true",
                   help="only rows whose snapshot data is presumptively stale")
    p.add_argument("--fields", help=f"comma list of {','.join(list(EXPECTED_HEADERS) + DERIVED_FIELDS)}")
    p.add_argument("--full", action="store_true", help="print all fields for each match")
    p.add_argument("--no-truncate", action="store_true")
    p.add_argument("--count-only", action="store_true")
    p.add_argument("--json", action="store_true", help="emit JSON_RESULTS= line")
    p.add_argument("--list-values", metavar="FIELD", help="distinct values of a field, with counts")
    p.add_argument("--no-freshness", action="store_true", help="skip the network freshness check")
    args = p.parse_args()

    path = find_data_file(args.data)
    snap, rows = load_rows(path)
    rows = [enrich(r) for r in rows]

    if args.list_values:
        key = args.list_values.upper()
        if key not in list(EXPECTED_HEADERS) + DERIVED_FIELDS:
            sys.exit(f"ERROR: unknown field {key}")
        from collections import Counter
        c = Counter()
        for r in rows:
            for part in re.split(r";\s*|\n", r.get(key, "")):
                if part.strip():
                    c[part.strip()] += 1
        for val, n in c.most_common():
            print(f"{n:5d}  {val}")
        print()
        print(footer(snap, len(rows), args))
        return

    matches = [r for r in rows if match(r, args)]

    if args.count_only:
        print(f"MATCHES: {len(matches)} of {len(rows)}")
    elif args.full:
        for r in matches:
            print("=" * 70)
            for key in list(EXPECTED_HEADERS) + DERIVED_FIELDS:
                v = r.get(key, "")
                if not args.no_truncate and len(v) > 500:
                    v = v[:500] + " …[truncated; use --no-truncate]"
                print(f"{key:14s}: {v}")
    else:
        fields = ([f.strip().upper() for f in args.fields.split(",")]
                  if args.fields else DEFAULT_FIELDS)
        bad = [f for f in fields if f not in list(EXPECTED_HEADERS) + DERIVED_FIELDS]
        if bad:
            sys.exit(f"ERROR: unknown field(s): {bad}")
        print(" | ".join(fields))
        for r in matches:
            vals = []
            for f in fields:
                v = r.get(f, "").replace("\n", " / ")
                if not args.no_truncate and len(v) > 160:
                    v = v[:160] + "…"
                vals.append(v)
            print(" | ".join(vals))
        print(f"\nMATCHES: {len(matches)} of {len(rows)}")

    if args.json:
        print("JSON_RESULTS=" + json.dumps(matches, ensure_ascii=False))

    print()
    print(footer(snap, len(rows), args))


if __name__ == "__main__":
    main()
